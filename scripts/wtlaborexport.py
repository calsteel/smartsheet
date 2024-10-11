import pyodbc
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import os

# Updated DSN connection with corrected parameters
DSN_NAME = "MASCSS"

# Connect to the database using the DSN
cnxn = pyodbc.connect(
    f"DSN={DSN_NAME};"
    "UID=scanco;"
    "PWD=;"
    r"Directory=\\mas2023\Sage\Sage 100\MAS90;"
    r"Prefix=\\mas2023\Sage\Sage 100\MAS90\SY\,\\mas2023\Sage\Sage 100\MAS90\==\;"
    r"ViewDLL=\\mas2023\Sage\Sage 100\MAS90\HOME;"
    "Company=CSS;"
    "SID=446;"
    r"LogFile=PVXODBC.LOG;"
    "CacheSize=4;"
    "DirtyReads=1;"
    "BurstMode=1;"
    "StripTrailingSpaces=1;",
    autocommit=True
)

cursor = cnxn.cursor()

# SQL query to select all rows from the table PM_WorkTicketHeader and PM_WorkTicketStep
detail_query = """
SELECT PM_LaborHistory.WorkTicketKey, PM_Employee.EmployeeNo, PM_Employee.FirstName, PM_Employee.LastName, 
PM_LaborHistory.StepLineKey, PM_LaborHistory.SequenceNo, PM_LaborHistory.TransactionDate, 
PM_LaborHistory.StartTime, PM_LaborHistory.EndDate, PM_LaborHistory.EndTime, PM_LaborHistory.RegisterNo, 
PM_LaborHistory.Overtime, PM_LaborHistory.StatusCode, PM_LaborHistory.StatusComment, 
PM_LaborHistory.HoursWorked, PM_LaborHistory.QuantityCompleted, PM_LaborHistory.DateCreated, PM_LaborHistory.TimeCreated
FROM {oj PM_LaborHistory PM_LaborHistory LEFT OUTER JOIN PM_Employee PM_Employee ON PM_LaborHistory.EmployeeKey = PM_Employee.EmployeeKey}
"""
cursor.execute(detail_query)

# Fetch the results from the query
detail_rows = cursor.fetchall()

# Specify the XLSX file path for the output
DETAILS_XLSX = "wtlaborexport.xlsx"

# If the file exists, load it, otherwise create a new workbook
if os.path.exists(DETAILS_XLSX):
    workbook = load_workbook(DETAILS_XLSX)
    sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active
    # Write the header row
    header = [column[0] for column in cursor.description]
    header.extend(["PO - PART # Helper", "comment", "Uploaded", "UniqueRows"])  # Add "Uploaded" and "UniqueRows" columns
    sheet.append(header)

# Create a set to track existing unique rows in the sheet
existing_unique_rows = {row[-1] for row in sheet.iter_rows(min_row=2, values_only=True)}

# Modify data rows and add them to the sheet
for row in detail_rows:
    # Generate UniqueRows value by combining DateCreated and TimeCreated
    unique_rows_value = f"{row.DateCreated}_{row.TimeCreated}"

    # Check if this UniqueRows value already exists in the sheet
    if unique_rows_value in existing_unique_rows:
        print(f"Row with UniqueRows value {unique_rows_value} already exists. Skipping.")
        continue

    # Process the new row
    work_ticket_no = str(row.WorkTicketKey).lstrip('0')
    step_no = str(row.StepLineKey).lstrip('0')

    # Convert StartTime
    decimal_time = float(row.StartTime)
    hours = int(decimal_time)
    minutes = int((decimal_time - hours) * 60)
    time_delta = timedelta(hours=hours, minutes=minutes)
    reference = datetime(2023, 1, 1)
    converted_time = reference + time_delta
    formatted_starttime = converted_time.strftime('%I:%M %p')

    # Convert EndTime
    enddecimal_time = float(row.EndTime)
    hours1 = int(enddecimal_time)
    minutes1 = int((enddecimal_time - hours1) * 60)
    time_delta1 = timedelta(hours=hours1, minutes=minutes1)
    reference1 = datetime(2023, 1, 1)
    converted_time1 = reference1 + time_delta1
    formatted_endtime = converted_time1.strftime('%I:%M %p')

    # Generate helper fields
    po_part_helper = f"{work_ticket_no} - {step_no}"
    comment = f"{row.TransactionDate} {formatted_starttime} to {row.EndDate} {formatted_endtime} : {row.FirstName} {row.LastName} - Quantity Complete: {int(row.QuantityCompleted)} , Status Code: {row.StatusCode} , Status Comment: {row.StatusComment}"

    # Convert the row to a list and append the new column values
    modified_row = list(row)
    modified_row.extend([po_part_helper, comment, "No", unique_rows_value])  # Default to "No" for Uploaded and add UniqueRows
    modified_row[0] = work_ticket_no  # Update WorkTicketNo
    modified_row[4] = step_no  # Update StepNo

    # Append the modified row to the sheet
    sheet.append(modified_row)
    existing_unique_rows.add(unique_rows_value)  # Add the new UniqueRows value to the set

# Save the workbook to the XLSX file
workbook.save(DETAILS_XLSX)

print(f"Data exported to {DETAILS_XLSX}")
